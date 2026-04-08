import io
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from constants import (
    TEMPLATE_PATH, AM_ROOMS, PM_ROOMS, AM_SHEET_TIMES, PM_SHEET_TIMES,
    SLOTS, ROWS, ROOM_START_COLS,
)

# ── Time helpers ──────────────────────────────────────────────────────────────

def to_minutes(time_str):
    """Convert a 'HH:MM' string to total minutes since midnight."""
    if not time_str:
        return 0
    hours, minutes = time_str.split(':')
    return int(hours) * 60 + int(minutes)

def covers_slot(time_ranges, room_time):
    """Return True if any of the person's time ranges covers the entire 30-minute slot."""
    room_minutes = to_minutes(room_time)
    return any(
        to_minutes(time_range['start']) <= room_minutes and to_minutes(time_range['end']) >= room_minutes + 30
        for time_range in time_ranges
    )

def person_on_sheet(person, is_pm):
    """Return True if the person has any availability overlapping the given sheet's time domain."""
    domain_times = PM_SHEET_TIMES if is_pm else AM_SHEET_TIMES
    domain_start = to_minutes(domain_times[0])
    domain_end   = to_minutes(domain_times[-1]) + 30
    return any(
        to_minutes(time_range['start']) < domain_end and to_minutes(time_range['end']) > domain_start
        for time_range in person['ranges']
    )

def safe_cell_value(value):
    """Prevent Excel formula injection by prefixing formula characters with a single quote."""
    s = str(value)
    return "'" + s if s and s[0] in ('=', '+', '-', '@', '\t', '\r') else s

def safe_excel_value(value):
    """Return a workbook-safe value while preserving numeric types when possible."""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return value
    return safe_cell_value(value)

def count_brothers_in_room(sorted_people, room_time):
    """Count brothers assigned to a room slot, including the Off row.

    A brother counts if they fully cover the room's 30-minute slot.
    The Off row always counts as 1.
    """
    brothers_from_people = sum(1 for person in sorted_people if covers_slot(person['ranges'], room_time))
    return brothers_from_people + 1  # +1 for the Off row

# ── Veil recommendation ───────────────────────────────────────────────────────

def recommend_veils(b, s, workers):
    """Recommend veil split to achieve equal rotations for brothers and sisters.

    Equal rotations means B/bv == S/sv, i.e. bv/sv == B/S.
    Finds the best split within worker constraints, up to 8 veils total.
    Brother veil = 2 workers, sister veil = 1 worker.
    """
    try:
        b       = int(b) if b else 0
        s       = int(s) if s else 0
        workers = int(workers) if workers else 0
    except (ValueError, TypeError):
        return None, None

    if workers == 0:
        return None, None

    if b == 0 and s == 0:
        return None, None

    best_bv    = 0
    best_sv    = 0
    best_diff  = float('inf')
    best_total = 0

    for bv in range(0, 9):
        for sv in range(0, 9 - bv):
            if bv + sv == 0:
                continue
            if bv + sv > 8:
                continue
            if bv * 2 + sv > workers:
                continue
            if bv > b:
                continue
            if sv > s:
                continue
            rot_b = b / bv if bv > 0 else float('inf')
            rot_s = s / sv if sv > 0 else float('inf')
            diff  = abs(rot_b - rot_s)
            total = bv + sv
            if diff < best_diff - 0.5 or (diff <= best_diff + 0.5 and total > best_total):
                best_diff  = diff
                best_bv    = bv
                best_sv    = sv
                best_total = total

    return best_bv, best_sv

# ── Build xlsx from template ──────────────────────────────────────────────────

def build_xlsx(people, room_data=None):
    """Fill the Excel template with people data and return a BytesIO buffer."""
    if room_data is None:
        room_data = {}
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template file not found: {TEMPLATE_PATH}")

    workbook = load_workbook(TEMPLATE_PATH)

    for sheet_name in workbook.sheetnames:
        worksheet      = workbook[sheet_name]
        is_pm          = 'PM' in sheet_name.upper()
        rooms          = PM_ROOMS if is_pm else AM_ROOMS
        visible_people = [person for person in people if person_on_sheet(person, is_pm)]
        sorted_people  = sorted(visible_people, key=lambda person: person['name'])

        # Write people data into rows 6-24
        for row_index in range(ROWS):
            is_off_row = (row_index == 0)
            person     = None if is_off_row else (sorted_people[row_index - 1] if row_index - 1 < len(sorted_people) else None)
            sheet_row  = 6 + row_index

            worksheet.cell(row=sheet_row, column=2).value = row_index + 1
            worksheet.cell(row=sheet_row, column=3).value = 'Off.' if is_off_row else (safe_cell_value(person['name']) if person else '')

            if person:
                for room_index, room in enumerate(rooms):
                    is_available = covers_slot(person['ranges'], room['time'])
                    if not is_available:
                        start_col = ROOM_START_COLS[room_index]
                        gray_fill = PatternFill(patternType='solid', fgColor='EDEDED')
                        for slot_index in range(SLOTS):
                            worksheet.cell(row=sheet_row, column=start_col + slot_index).fill = gray_fill

        # Write brothers count into row 26 for each room
        for room_index, room in enumerate(rooms):
            start_col = ROOM_START_COLS[room_index]
            count     = count_brothers_in_room(sorted_people, room['time'])
            worksheet.cell(row=26, column=start_col).value = count

        # Write officiator, B:, S: into header cells for each room
        for room_index, room in enumerate(rooms):
            start_col = ROOM_START_COLS[room_index]
            rd        = room_data.get(room['time'], {})
            if rd.get('off'):
                off_cell       = worksheet.cell(row=4, column=start_col + 3)
                off_cell.value = safe_cell_value(rd['off'])
                off_cell.font  = Font(bold=True, color='335593')
            b_val   = rd.get('b', '')
            s_val   = rd.get('s', '')
            workers = count_brothers_in_room(sorted_people, room['time'])
            if b_val:
                worksheet.cell(row=3, column=start_col + 2).value = safe_excel_value(b_val)
            if s_val:
                worksheet.cell(row=3, column=start_col + 8).value = safe_excel_value(s_val)
            if b_val or s_val:
                bv, sv = recommend_veils(b_val, s_val, workers)
                if bv is not None:
                    worksheet.cell(row=3, column=start_col + 4).value = f"{bv}B"
                if sv is not None:
                    worksheet.cell(row=3, column=start_col + 10).value = f"{sv}S"

    # ── Source data sheet ─────────────────────────────────────────────────────
    # Each row is a single paste-ready line matching the paste box format:
    #   Person:  Name, start, end[, start, end ...]
    #   Room:    time, officiator, b, s
    src = workbook.create_sheet(title='Source Data')

    row = 1
    for person in sorted(people, key=lambda p: p['name']):
        parts = [safe_cell_value(person['name'])] + [safe_cell_value(t) for r in person['ranges'] for t in (r['start'], r['end'])]
        src.cell(row=row, column=1).value = ', '.join(parts)
        row += 1

    row += 1  # blank separator

    all_times = [r['time'] for r in AM_ROOMS] + [r['time'] for r in PM_ROOMS]
    for time in all_times:
        rd = room_data.get(time, {})
        if not any([rd.get('off'), rd.get('b'), rd.get('s')]):
            continue
        parts = [
            safe_cell_value(time),
            safe_cell_value(rd.get('off', '')),
            safe_cell_value(rd.get('b', '')),
            safe_cell_value(rd.get('s', '')),
        ]
        src.cell(row=row, column=1).value = ', '.join(parts)
        row += 1

    max_len = max((len(str(c.value)) for c in src['A'] if c.value), default=20)
    src.column_dimensions['A'].width = min(max_len + 4, 80)

    xlsx_buffer = io.BytesIO()
    workbook.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    return xlsx_buffer
