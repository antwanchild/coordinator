import io
import os
import time
import platform
import logging
from logging.handlers import TimedRotatingFileHandler
from datetime import datetime

from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 1 * 1024 * 1024  # 1MB max request size

TEMPLATE_PATH = "V-COORDINATE--Scheduled.xlsx"
APP_VERSION   = os.environ.get('APP_VERSION', 'dev')

# ── Logging ───────────────────────────────────────────────────────────────────

log_formatter = logging.Formatter(
    '%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S %Z'
)
logger = logging.getLogger('coordinator')
LOG_LEVEL = os.environ.get('LOG_LEVEL', 'INFO').upper()
logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))

# Write to /config/logs if the volume is mounted, otherwise console only
LOG_DIR = '/config/logs' if os.path.isdir('/config') else None

if LOG_DIR:
    os.makedirs(LOG_DIR, exist_ok=True)

    class WeeklyLogHandler(TimedRotatingFileHandler):
        """Rotate logs weekly, keeping .log extension on rotated files."""
        def rotation_filename(self, default_name):
            base   = os.path.join(LOG_DIR, 'coordinator')
            suffix = default_name.split('.')[-1]
            return f"{base}.{suffix}.log"

    file_handler = WeeklyLogHandler(
        os.path.join(LOG_DIR, 'coordinator.log'),
        when='W0', interval=1, backupCount=4, utc=False
    )
    file_handler.setFormatter(log_formatter)
    logger.addHandler(file_handler)

stream_handler = logging.StreamHandler()
stream_handler.setFormatter(log_formatter)
logger.addHandler(stream_handler)

# ── Request logging ───────────────────────────────────────────────────────────

# Suppress Werkzeug's default request logger
logging.getLogger('werkzeug').setLevel(logging.ERROR)

ROUTE_LABELS = {
    '/':        'App loaded',
    '/preview': 'Preview requested',
    '/export':  'Export requested',
}

@app.before_request
def before_request():
    request.start_time = time.monotonic()

@app.after_request
def after_request(response):
    duration_ms = int((time.monotonic() - request.start_time) * 1000)
    label = ROUTE_LABELS.get(request.path, request.path)
    logger.info(f"{label} | {response.status_code} ({duration_ms}ms)")
    return response

# ── Layout constants ──────────────────────────────────────────────────────────

def chars_to_pixels(char_width):
    """Convert Excel character-width units to pixels (approximate)."""
    return max(4, int(char_width * 7 + 5))

def points_to_pixels(point_height):
    """Convert Excel point height to pixels at 96 DPI."""
    return max(4, int(point_height * 96 / 72))

# Column widths in Excel character units (used only as fallback reference)
COL_WIDTHS_CHARS = {1: 5.28, 2: 3.85, 3: 11.28, 4: 1.85}
for col in range(5, 66):
    COL_WIDTHS_CHARS[col] = 13.0

# Row heights in Excel point units
ROW_HEIGHTS_PT = {1: 17.25, 2: 15.0, 3: 15.0, 4: 15.75, 5: 20.25}
for row in range(6, 24):
    ROW_HEIGHTS_PT[row] = 18.0
for row in range(24, 31):
    ROW_HEIGHTS_PT[row] = 15.0
ROW_HEIGHTS_PT[30] = 15.75

# Pixel widths per column — overridden for key columns to match spreadsheet layout
COL_PX = {}
COL_PX[1] = 10   # A: spacer
COL_PX[2] = 31   # B: row numbers
COL_PX[3] = 100  # C: names
for col in range(4, 66):
    COL_PX[col] = 24  # D onward: uniform slot columns

# Pixel heights per row, derived from point heights
ROW_PX = {row: points_to_pixels(height) for row, height in ROW_HEIGHTS_PT.items()}

SLOTS           = 12                      # number of slot columns per room block
ROWS            = 18                      # number of data rows (row 0 = Off, rows 1-17 = people)
ROOM_START_COLS = [4, 16, 28, 40, 52]    # first column of each room block

# ── Room definitions ──────────────────────────────────────────────────────────

AM_ROOMS = [
    {'label': 'Room: 1', 'time': '11:00', 'time_raw': '  Time: 11:00+', 'n24': '#5--tight fit to Cel Room', 'n25': '',                        'red': True,  'letter': 'B', 'wlkr': '(#2/#3)'},
    {'label': 'Room: 2', 'time': '11:30', 'time_raw': '  Time: 11:30+', 'n24': '',                          'n25': '',                        'red': False, 'letter': '',  'wlkr': ''},
    {'label': 'Room: 3', 'time': '12:00', 'time_raw': '  Time: 12:00+', 'n24': '',                          'n25': '',                        'red': False, 'letter': '',  'wlkr': ''},
    {'label': 'Room: 4', 'time': '12:30', 'time_raw': '  Time: 12:30+', 'n24': '#4--tight fit to Cel Room', 'n25': 'Use thin rec. for #1-#4', 'red': True,  'letter': 'S', 'wlkr': '(#6/#7)'},
    {'label': 'Room: 1', 'time': '13:00', 'time_raw': '  Time: 1:00+',  'n24': '#5--tight fit to Cel Room', 'n25': '',                        'red': True,  'letter': 'B', 'wlkr': '(#2/#3)'},
]

PM_ROOMS = [
    {'label': 'Room: 3', 'time': '14:00', 'time_raw': '  Time: 2:00+', 'n24': '',                          'n25': '',                        'red': False, 'letter': '',  'wlkr': ''},
    {'label': 'Room: 4', 'time': '14:30', 'time_raw': '  Time: 2:30+', 'n24': '#4--tight fit to Cel Room', 'n25': 'Use thin rec. for #1-#4', 'red': True,  'letter': 'S', 'wlkr': '(#6/#7)'},
    {'label': 'Room: 1', 'time': '15:00', 'time_raw': '  Time: 3:00+', 'n24': '#5--tight fit to Cel Room', 'n25': '',                        'red': True,  'letter': 'B', 'wlkr': '(#2/#3)'},
    {'label': 'Room: 2', 'time': '15:30', 'time_raw': '  Time: 3:30+', 'n24': '',                          'n25': '',                        'red': False, 'letter': '',  'wlkr': ''},
    {'label': 'Room: 3', 'time': '16:00', 'time_raw': '  Time: 4:00+', 'n24': '',                          'n25': '',                        'red': False, 'letter': '',  'wlkr': ''},
]

# Slot colors and labels cycle across the 12 slot columns in each room block
SLOT_COLORS = [
    '#B4C6E7', '#FFFFFF', '#FFE598', '#C5DEB5',
    '#B4C6E7', '#B4C6E7',
    '#FFFFFF', '#FFFFFF',
    '#FFE598', '#FFE598',
    '#C5DEB5', '#C5DEB5',
]
SLOT_LABELS       = ['1', '2', '3', '4', '5P', '5', '6P', '6', '7P', '7', '8P', '8']
UNAVAILABLE_COLOR = '#FFFF00'  # yellow fill for slots where person is unavailable

# ── Time helpers ──────────────────────────────────────────────────────────────

def to_minutes(time_str):
    """Convert a 'HH:MM' string to total minutes since midnight."""
    if not time_str:
        return 0
    hours, minutes = time_str.split(':')
    return int(hours) * 60 + int(minutes)

def covers_slot(time_ranges, room_time):
    """Return True if any of the person's time ranges covers the entire 30-minute slot."""
    room_minutes = to_minutes(room_time)
    return any(
        to_minutes(time_range['start']) <= room_minutes and to_minutes(time_range['end']) >= room_minutes + 30
        for time_range in time_ranges
    )

def person_on_sheet(person, is_pm):
    """Return True if the person has any availability overlapping the given sheet's time domain."""
    domain_times = ['14:00', '14:30', '15:00', '15:30', '16:00'] if is_pm else ['11:00', '11:30', '12:00', '12:30', '13:00']
    domain_start = to_minutes(domain_times[0])
    domain_end   = to_minutes(domain_times[-1]) + 30
    return any(
        to_minutes(time_range['start']) < domain_end and to_minutes(time_range['end']) > domain_start
        for time_range in person['ranges']
    )

def count_brothers_in_room(sorted_people, room_time):
    """Count brothers assigned to a room slot, including the Off row.

    A brother counts if they fully cover the room's 30-minute slot.
    The Off row always counts as 1.
    """
    brothers_from_people = sum(1 for person in sorted_people if covers_slot(person['ranges'], room_time))
    return brothers_from_people + 1  # +1 for the Off row

def brothers_label(count):
    """Return a formatted string like '4 Brothers' or '1 Brother'."""
    return f"{count} {'Brother' if count == 1 else 'Brothers'}"

# ── Build xlsx from template ──────────────────────────────────────────────────

def build_xlsx(people):
    """Fill the Excel template with people data and return a BytesIO buffer."""
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template file not found: {TEMPLATE_PATH}")

    workbook = load_workbook(TEMPLATE_PATH)

    for sheet_name in workbook.sheetnames:
        worksheet      = workbook[sheet_name]
        is_pm          = 'PM' in sheet_name.upper()
        rooms          = PM_ROOMS if is_pm else AM_ROOMS
        visible_people = [person for person in people if person_on_sheet(person, is_pm)]
        sorted_people  = sorted(visible_people, key=lambda person: person['name'])

        # Write people data into rows 6-23
        for row_index in range(ROWS):
            is_off_row = (row_index == 0)
            person     = None if is_off_row else (sorted_people[row_index - 1] if row_index - 1 < len(sorted_people) else None)
            sheet_row  = 6 + row_index

            worksheet.cell(row=sheet_row, column=2).value = row_index + 1
            worksheet.cell(row=sheet_row, column=3).value = 'Off.' if is_off_row else (person['name'] if person else '')

            if person:
                for room_index, room in enumerate(rooms):
                    is_available = covers_slot(person['ranges'], room['time'])
                    if not is_available:
                        start_col   = ROOM_START_COLS[room_index]
                        yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')
                        for slot_index in range(SLOTS):
                            worksheet.cell(row=sheet_row, column=start_col + slot_index).fill = yellow_fill

        # Write brothers count into row 25 (# Bro's Avail) for each room
        for room_index, room in enumerate(rooms):
            start_col = ROOM_START_COLS[room_index]
            count     = count_brothers_in_room(sorted_people, room['time'])
            worksheet.cell(row=25, column=start_col).value = count

    xlsx_buffer = io.BytesIO()
    workbook.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    return xlsx_buffer

# ── Render preview image ──────────────────────────────────────────────────────

def hex_to_rgb(hex_color):
    """Convert a hex color string like '#RRGGBB' to an (R, G, B) tuple."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))

def render_preview(people, is_pm):
    """Render the schedule as a PNG image and return a BytesIO buffer."""
    rooms          = PM_ROOMS if is_pm else AM_ROOMS
    visible_people = [person for person in people if person_on_sheet(person, is_pm)]
    sorted_people  = sorted(visible_people, key=lambda person: person['name'])

    # Build cumulative x positions for each column
    x_positions = {}
    x = 0
    for col in range(1, 66):
        x_positions[col] = x
        x += COL_PX.get(col, chars_to_pixels(13.0))
    total_width = x

    # Build cumulative y positions for each row
    y_positions = {}
    y = 0
    for row in range(1, 31):
        y_positions[row] = y
        y += ROW_PX.get(row, points_to_pixels(15.0))
    total_height = y

    image = Image.new('RGB', (total_width, total_height), 'white')
    draw  = ImageDraw.Draw(image)

    # Load fonts — fall back to default if DejaVu is not available
    try:
        font_small  = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 9)
        font_medium = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 11)
        font_bold   = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 11)
        font_bold14 = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 14)
        font_bold18 = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 18)
    except Exception:
        font_small = font_medium = font_bold = font_bold14 = font_bold18 = ImageFont.load_default()

    # ── Drawing helpers ───────────────────────────────────────────────────────

    def cell_rect(col, row, colspan=1, rowspan=1):
        """Return the (x1, y1, x2, y2) pixel bounds of a cell or merged range."""
        x1 = x_positions[col]
        y1 = y_positions[row]
        x2 = x_positions[col + colspan] if (col + colspan) in x_positions else total_width
        y2 = y_positions[row + rowspan] if (row + rowspan) in y_positions else total_height
        return x1, y1, x2, y2

    def fill_cell(col, row, color, colspan=1, rowspan=1):
        """Fill a cell or merged range with a solid color."""
        x1, y1, x2, y2 = cell_rect(col, row, colspan, rowspan)
        draw.rectangle([x1, y1, x2 - 1, y2 - 1], fill=hex_to_rgb(color))

    def draw_text(col, row, text, font, color='#000000', align='left', colspan=1, rowspan=1):
        """Draw text inside a cell, truncating if it overflows."""
        x1, y1, x2, y2 = cell_rect(col, row, colspan, rowspan)
        rgb     = hex_to_rgb(color)
        padding = 3

        # Truncate text to fit within cell width
        while text and draw.textlength(text, font=font) > (x2 - x1 - padding * 2):
            text = text[:-1]

        text_width  = draw.textlength(text, font=font)
        cell_height = y2 - y1

        if align == 'center':
            tx = x1 + (x2 - x1 - text_width) // 2
        elif align == 'right':
            tx = x2 - text_width - padding
        else:
            tx = x1 + padding

        font_height = (font.getbbox(text)[3] - font.getbbox(text)[1]) if text else 10
        ty = y1 + (cell_height - font_height) // 2
        draw.text((tx, ty), text, fill=rgb, font=font)

    def draw_border(col, row, colspan=1, rowspan=1, left=None, right=None, top=None, bottom=None):
        """Draw borders on one or more sides of a cell range.

        Border style options: 'double', 'medium', 'thin'
        """
        x1, y1, x2, y2 = cell_rect(col, row, colspan, rowspan)
        x2 -= 1
        y2 -= 1

        def border_style(style):
            """Map a border style name to (color_rgb, line_width)."""
            if style in ('double', 'medium'):
                return hex_to_rgb('#595959'), 2
            if style == 'thin':
                return hex_to_rgb('#aaaaaa'), 1
            return None

        if left:
            color, width = border_style(left)
            draw.line([(x1, y1), (x1, y2)], fill=color, width=width)
        if right:
            color, width = border_style(right)
            draw.line([(x2, y1), (x2, y2)], fill=color, width=width)
        if top:
            color, width = border_style(top)
            draw.line([(x1, y1), (x2, y1)], fill=color, width=width)
        if bottom:
            color, width = border_style(bottom)
            draw.line([(x1, y2), (x2, y2)], fill=color, width=width)

    # ── Header rows 1–5 ───────────────────────────────────────────────────────

    sheet_label = 'P.M.' if is_pm else 'A.M.'
    fill_cell(3, 1, '#FFFFFF', colspan=1, rowspan=4)
    draw_text(3, 1, sheet_label, font_bold18, color='#FF0000', align='center', colspan=1, rowspan=4)
    draw_border(3, 1, colspan=1, rowspan=4, left='medium', right='medium', top='double', bottom='double')

    for room_index, room in enumerate(rooms):
        start_col    = ROOM_START_COLS[room_index]
        is_last_room = (room_index == len(rooms) - 1)

        # Row 1: Room label — full width, "Room: " black, number red
        fill_cell(start_col, 1, '#FFFFFF', colspan=SLOTS)
        room_prefix      = room['label'].split(':')[0] + ': '
        room_number      = room['label'].split(': ')[1] if ': ' in room['label'] else ''
        x1, y1, x2, y2  = cell_rect(start_col, 1, SLOTS, 1)
        prefix_width     = draw.textlength(room_prefix, font=font_bold14)
        total_text_width = prefix_width + draw.textlength(room_number, font=font_bold14)
        tx = x1 + (x2 - x1 - total_text_width) // 2
        ty = y1 + (y2 - y1 - font_bold14.getbbox('A')[3]) // 2
        draw.text((tx, ty), room_prefix, fill=hex_to_rgb('#000000'), font=font_bold14)
        draw.text((tx + int(prefix_width), ty), room_number, fill=hex_to_rgb('#FF0000'), font=font_bold14)
        draw_border(start_col, 1, colspan=SLOTS, left='medium', top='double', bottom='thin',
                    right='medium' if is_last_room else 'thin')

        # Row 2: "Time:" in left merge, time value in red in right merge
        fill_cell(start_col, 2, '#FFFFFF', colspan=SLOTS)
        time_parts  = room['time_raw'].strip().split(': ', 1)
        time_prefix = time_parts[0] + ': ' if len(time_parts) > 1 else room['time_raw'].strip()
        time_value  = time_parts[1] if len(time_parts) > 1 else ''
        # "Time:" centered in left 2 cols, time value left-aligned in remaining cols
        x1l, y1l, x2l, y2l = cell_rect(start_col, 2, 2, 1)
        x1r, y1r, x2r, y2r = cell_rect(start_col + 2, 2, SLOTS - 2, 1)
        ty = y1l + (y2l - y1l - font_bold.getbbox('A')[3]) // 2
        prefix_w = draw.textlength(time_prefix, font=font_bold)
        draw.text((x1l + (x2l - x1l - prefix_w) // 2, ty), time_prefix, fill=hex_to_rgb('#000000'), font=font_bold)
        draw.text((x1r + 3, ty), time_value, fill=hex_to_rgb('#FF0000'), font=font_bold)
        draw_border(start_col, 2, colspan=SLOTS, left='medium', top='thin', bottom='thin',
                    right='medium' if is_last_room else 'thin')

        # Row 3: B: left half, S: right half
        half_slots = SLOTS // 2
        fill_cell(start_col, 3, '#FFFFFF', colspan=half_slots)
        draw_text(start_col + 1, 3, 'B:', font_bold, colspan=half_slots - 1)
        draw_border(start_col, 3, colspan=half_slots, left='medium', top='thin', bottom='thin')
        fill_cell(start_col + half_slots, 3, '#FFFFFF', colspan=half_slots)
        draw_text(start_col + half_slots, 3, 'S:', font_bold, colspan=half_slots)
        draw_border(start_col + half_slots, 3, colspan=half_slots, top='thin', bottom='thin',
                    right='medium' if is_last_room else 'thin')

        # Row 4: Full width "Off:"
        fill_cell(start_col, 4, '#FFFFFF', colspan=SLOTS)
        draw_text(start_col + 1, 4, 'Off:', font_bold, colspan=SLOTS - 1)
        draw_border(start_col, 4, colspan=SLOTS, left='medium', top='thin', bottom='double',
                    right='medium' if is_last_room else 'thin')

        # Row 5: Slot number labels
        for slot_index, slot_label in enumerate(SLOT_LABELS):
            col          = start_col + slot_index
            is_last_slot = (slot_index == SLOTS - 1)
            fill_cell(col, 5, '#FFFFFF')
            draw_text(col, 5, slot_label, font_small, align='center')
            draw_border(col, 5,
                        left='medium' if slot_index == 0 else 'thin',
                        top='double',
                        bottom='thin',
                        right='medium' if (is_last_room and is_last_slot) else ('thin' if is_last_slot else None))

    draw_border(2, 5, left='thin', right='double', top='double', bottom='thin')
    draw_border(3, 5, left='double', top='double', bottom='thin')

    # ── Data rows 6–23 ────────────────────────────────────────────────────────

    for row_index in range(ROWS):
        sheet_row   = 6 + row_index
        is_off_row  = (row_index == 0)
        is_last_row = (row_index == ROWS - 1)
        person      = None if is_off_row else (sorted_people[row_index - 1] if row_index - 1 < len(sorted_people) else None)
        name_value  = 'Off.' if is_off_row else (person['name'] if person else '')

        # Column B: row number
        fill_cell(2, sheet_row, '#FFFFFF')
        draw_text(2, sheet_row, str(row_index + 1), font_medium, align='right')
        draw_border(2, sheet_row, left='thin', right='double', top='thin',
                    bottom='thin' if is_last_row else None)

        # Column C: person name
        fill_cell(3, sheet_row, '#FFFFFF')
        draw_text(3, sheet_row, name_value, font_medium, align='center')
        draw_border(3, sheet_row, left='double', right='medium', top='thin',
                    bottom='thin' if is_last_row else None)

        # Slot columns: fill with slot color or yellow if unavailable
        for room_index, room in enumerate(rooms):
            start_col    = ROOM_START_COLS[room_index]
            is_last_room = (room_index == len(rooms) - 1)
            is_available = person and covers_slot(person['ranges'], room['time'])

            for slot_index in range(SLOTS):
                col          = start_col + slot_index
                is_last_slot = (slot_index == SLOTS - 1)
                slot_color   = SLOT_COLORS[slot_index] if (is_off_row or not person or is_available) else UNAVAILABLE_COLOR
                fill_cell(col, sheet_row, slot_color)
                draw_border(col, sheet_row,
                            left='medium' if slot_index == 0 else 'thin',
                            top='thin',
                            bottom='thin' if is_last_row else None,
                            right='medium' if (is_last_room and is_last_slot) else ('thin' if is_last_slot else None))

    # ── Footer rows 24–30 ─────────────────────────────────────────────────────

    FOOTER_LABELS = ['', "# Bro's Avail", 'Narrow Side?', 'Live?', 'Wchr/Wlkr?', 'Language?', 'Other?']

    for footer_index, footer_label in enumerate(FOOTER_LABELS):
        sheet_row      = 24 + footer_index
        is_last_footer = (footer_index == len(FOOTER_LABELS) - 1)

        fill_cell(3, sheet_row, '#FFFFFF')
        if footer_label:
            draw_text(3, sheet_row, footer_label, font_small, align='center')
        draw_border(3, sheet_row, left='double', bottom='double' if is_last_footer else None)

        for room_index, room in enumerate(rooms):
            start_col    = ROOM_START_COLS[room_index]
            is_last_room = (room_index == len(rooms) - 1)

            if footer_index == 0:
                # First footer row: room note text
                fill_cell(start_col, sheet_row, '#FFFFFF', colspan=SLOTS)
                if room['n24']:
                    draw_text(start_col, sheet_row, room['n24'], font_small, colspan=SLOTS)
                draw_border(start_col, sheet_row, colspan=SLOTS, left='medium', top='thin',
                            right='medium' if is_last_room else 'thin')

            elif footer_index == 1:
                # Brothers count — single checkbox-sized box, rest of row empty
                room_count = count_brothers_in_room(sorted_people, room['time'])
                fill_cell(start_col, sheet_row, '#FFFFFF')
                draw_text(start_col, sheet_row, str(room_count), font_bold, align='center')
                draw_border(start_col, sheet_row, left='medium', right='thin', top='thin', bottom='thin')
                # Fill remaining cols with empty bordered cells
                fill_cell(start_col + 1, sheet_row, '#FFFFFF', colspan=SLOTS - 1)
                draw_border(start_col + 1, sheet_row, colspan=SLOTS - 1, left='thin',
                            right='medium' if is_last_room else 'thin')

            else:
                # Remaining footer rows: checkbox column + label column
                checkbox_color = '#FF0000' if (room['red'] and footer_index == 2) else '#FFFFFF'
                fill_cell(start_col, sheet_row, checkbox_color)
                draw_border(start_col, sheet_row, left='medium', right='thin', top='thin',
                            bottom='thin' if not is_last_footer else 'double')

                fill_cell(start_col + 1, sheet_row, '#FFFFFF', colspan=SLOTS - 1)
                if footer_index == 2 and room['letter']:
                    draw_text(start_col + 1, sheet_row, room['letter'], font_bold, colspan=SLOTS - 1)
                elif footer_index == 4 and room['wlkr']:
                    draw_text(start_col + 1, sheet_row, room['wlkr'], font_bold, colspan=SLOTS - 1)
                draw_border(start_col + 1, sheet_row, colspan=SLOTS - 1, left='thin',
                            right='medium' if is_last_room else 'thin',
                            bottom='double' if is_last_footer else None)

    # ── Column B borders (data rows) ──────────────────────────────────────────
    for row_index in range(ROWS):
        draw_border(2, 6 + row_index, left='thin', right='double', top='thin')

    image_buffer = io.BytesIO()
    image.save(image_buffer, format='PNG', optimize=True)
    image_buffer.seek(0)
    return image_buffer

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', version=APP_VERSION, now=datetime.now())


@app.route('/preview', methods=['POST'])
def preview():
    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        logger.warning("Preview request | invalid or missing JSON body")
        return jsonify({'error': 'Invalid request body'}), 400

    people = data.get('people', [])
    if not isinstance(people, list):
        return jsonify({'error': 'people must be a list'}), 400

    is_pm      = bool(data.get('is_pm', False))
    sheet_name = 'PM' if is_pm else 'AM'

    try:
        image_buffer = render_preview(people, is_pm)

        rooms        = PM_ROOMS if is_pm else AM_ROOMS
        visible      = [p for p in people if person_on_sheet(p, is_pm)]
        sorted_ppl   = sorted(visible, key=lambda p: p['name'])
        brothers     = [count_brothers_in_room(sorted_ppl, room['time']) for room in rooms]

        logger.info(f"Preview generated | sheet={sheet_name}, people={len(visible)}, brothers={brothers}")
        return send_file(image_buffer, mimetype='image/png')
    except FileNotFoundError as e:
        logger.error(f"Template missing: {e}")
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        logger.error(f"Preview failed | sheet={sheet_name}, people={len(people)}, error={e}")
        return jsonify({'error': 'Preview generation failed'}), 500


@app.route('/export', methods=['POST'])
def export():
    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        logger.warning("Export request | invalid or missing JSON body")
        return jsonify({'error': 'Invalid request body'}), 400

    people = data.get('people', [])
    if not isinstance(people, list):
        return jsonify({'error': 'people must be a list'}), 400

    try:
        xlsx_buffer = build_xlsx(people)

        am_visible  = sorted([p for p in people if person_on_sheet(p, False)], key=lambda p: p['name'])
        pm_visible  = sorted([p for p in people if person_on_sheet(p, True)],  key=lambda p: p['name'])
        am_brothers = [count_brothers_in_room(am_visible, room['time']) for room in AM_ROOMS]
        pm_brothers = [count_brothers_in_room(pm_visible, room['time']) for room in PM_ROOMS]

        logger.info(f"Export generated | people={len(people)}, brothers=AM:{am_brothers} PM:{pm_brothers}")
        return send_file(
            xlsx_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='schedule-output.xlsx'
        )
    except FileNotFoundError as e:
        logger.error(f"Template missing: {e}")
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        logger.error(f"Export failed | people={len(people)}, error={e}")
        return jsonify({'error': 'Export generation failed'}), 500


if __name__ == '__main__':
    log_dir_display = LOG_DIR if LOG_DIR else 'console only'
    logger.info("=" * 60)
    logger.info(f"Coordinator v{APP_VERSION} starting on port 8080")
    logger.info(f"template={TEMPLATE_PATH} | logs={log_dir_display} | log_level={LOG_LEVEL}")
    logger.info(f"Python {platform.python_version()} | {platform.system()} {platform.release()} | {platform.machine()}")
    app.run(host='0.0.0.0', port=8080, debug=False)

