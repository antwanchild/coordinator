import os
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app
from schedule import (
    build_xlsx,
    count_brothers_in_room,
    covers_slot,
    person_on_sheet,
    recommend_veils,
)


def make_person(name, ranges):
    return {'name': name, 'ranges': ranges}


class ScheduleLogicTests(unittest.TestCase):
    def test_covers_slot_requires_full_coverage(self):
        self.assertTrue(covers_slot([{'start': '11:00', 'end': '11:30'}], '11:00'))
        self.assertFalse(covers_slot([{'start': '11:00', 'end': '11:15'}], '11:00'))
        self.assertFalse(covers_slot([{'start': '11:30', 'end': '12:00'}], '11:00'))

    def test_person_on_sheet_filters_am_and_pm_ranges(self):
        person = make_person('Alex', [{'start': '11:00', 'end': '11:30'}])
        self.assertTrue(person_on_sheet(person, False))
        self.assertFalse(person_on_sheet(person, True))

        person = make_person('Blake', [{'start': '14:00', 'end': '14:30'}])
        self.assertFalse(person_on_sheet(person, False))
        self.assertTrue(person_on_sheet(person, True))

    def test_count_brothers_in_room_includes_off_row(self):
        people = [
            make_person('Alex', [{'start': '11:00', 'end': '11:30'}]),
            make_person('Blake', [{'start': '11:30', 'end': '12:00'}]),
            make_person('Casey', [{'start': '12:00', 'end': '12:30'}]),
        ]
        self.assertEqual(count_brothers_in_room(people, '11:00'), 2)
        self.assertEqual(count_brothers_in_room(people, '11:30'), 2)
        self.assertEqual(count_brothers_in_room(people, '12:00'), 2)

    def test_recommend_veils_prefers_balanced_split_within_limits(self):
        self.assertEqual(recommend_veils(4, 4, 6), (2, 2))
        self.assertEqual(recommend_veils(8, 2, 6), (2, 1))
        self.assertEqual(recommend_veils(0, 0, 6), (None, None))

    def test_build_xlsx_writes_source_sheet_and_values(self):
        people = [
            make_person('Alex', [{'start': '11:00', 'end': '11:30'}]),
            make_person('Blake', [{'start': '14:00', 'end': '14:30'}]),
        ]
        room_data = {
            '11:00': {'off': 'Officer A', 'b': '3', 's': '3'},
            '14:00': {'off': 'Officer B', 'b': '2', 's': '1'},
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, 'template.xlsx')
            workbook = Workbook()
            workbook.remove(workbook.active)
            workbook.create_sheet('AM')
            workbook.create_sheet('PM')
            workbook.save(template_path)

            with patch('schedule.TEMPLATE_PATH', template_path):
                buffer = build_xlsx(people, room_data)

        result = load_workbook(buffer)
        self.assertIn('Source Data', result.sheetnames)

        source = result['Source Data']
        self.assertEqual(source['A1'].value, 'Alex, 11:00, 11:30')
        self.assertEqual(source['A2'].value, 'Blake, 14:00, 14:30')
        self.assertEqual(source['A4'].value, '11:00, Officer A, 3, 3')

    def test_validate_request_payload_normalizes_and_rejects_invalid_values(self):
        payload, error = app.validate_request_payload({
            'people': [{'name': '  Alex  ', 'ranges': [{'start': '11:00', 'end': '11:30'}]}],
            'room_data': {'11:00': {'off': '  Officer A  ', 'b': 2, 's': '1'}},
            'is_pm': True,
        }, require_is_pm=True)

        self.assertIsNone(error)
        assert payload is not None
        self.assertEqual(payload['people'][0]['name'], 'Alex')
        self.assertEqual(payload['room_data']['11:00']['off'], 'Officer A')
        self.assertEqual(payload['room_data']['11:00']['b'], '2')
        self.assertEqual(payload['room_data']['11:00']['s'], '1')
        self.assertTrue(payload['is_pm'])

        _, error = app.validate_request_payload({
            'people': [],
            'room_data': {},
            'is_pm': 'yes',
        }, require_is_pm=True)
        self.assertEqual(error, 'is_pm must be a boolean')

        _, error = app.validate_request_payload({
            'people': [{'name': 'Alex', 'ranges': [{'start': '11:00', 'end': '11:10'}]}],
            'room_data': {},
            'is_pm': False,
        })
        self.assertEqual(error, 'people[0].ranges[0] must use 15-minute times between 11:00 and 16:45')

    def test_health_route_reports_json(self):
        client = app.app.test_client()
        response = client.get('/health')

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['status'], 'ok')
        self.assertIn('version', payload)
        self.assertIn('commit', payload)


if __name__ == '__main__':
    unittest.main()
