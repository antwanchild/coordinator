import io
import os
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app
import schedule


class RouteTests(unittest.TestCase):
    def test_root_route_returns_html(self):
        client = app.app.test_client()
        response = client.get('/')

        self.assertEqual(response.status_code, 200)
        mimetype = response.mimetype or ''
        self.assertIn('text/html', mimetype)
        self.assertIn(b'V-COORDINATOR SCHEDULE BUILDER', response.data)

    def test_health_route_reports_json(self):
        client = app.app.test_client()
        response = client.get('/health')

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['status'], 'ok')
        self.assertIn('version', payload)
        self.assertIn('commit', payload)

    def test_preview_route_returns_png_for_valid_payload(self):
        client = app.app.test_client()

        with patch.object(app, 'render_preview', return_value=io.BytesIO(b'fake-png')):
            response = client.post('/preview', json={
                'people': [{'name': 'Alex', 'ranges': [{'start': '11:00', 'end': '11:30'}]}],
                'room_data': {},
                'is_pm': True,
            })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, 'image/png')
        self.assertEqual(response.data, b'fake-png')

    def test_preview_route_rejects_non_boolean_is_pm(self):
        client = app.app.test_client()

        response = client.post('/preview', json={
            'people': [{'name': 'Alex', 'ranges': [{'start': '11:00', 'end': '11:30'}]}],
            'room_data': {},
            'is_pm': 'yes',
        })

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertEqual(payload['error'], 'is_pm must be a boolean')

    def test_preview_route_rejects_malformed_json_body(self):
        client = app.app.test_client()

        response = client.post(
            '/preview',
            data='not-json',
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertEqual(payload['error'], 'Invalid request body')

    def test_export_route_rejects_invalid_room_data(self):
        client = app.app.test_client()

        response = client.post('/export', json={
            'people': [{'name': 'Alex', 'ranges': [{'start': '11:00', 'end': '11:30'}]}],
            'room_data': {'10:00': {'off': 'Officer A'}},
            'is_pm': False,
        })

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertEqual(payload['error'], 'room_data contains unsupported time "10:00"')

    def test_export_route_rejects_empty_body(self):
        client = app.app.test_client()

        response = client.post('/export', data='', content_type='application/json')

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertEqual(payload['error'], 'Invalid request body')

    def test_export_route_returns_xlsx_for_valid_payload(self):
        client = app.app.test_client()

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, 'template.xlsx')
            workbook = Workbook()
            active_sheet = workbook.active
            if active_sheet is not None:
                workbook.remove(active_sheet)
            workbook.create_sheet('AM')
            workbook.create_sheet('PM')
            workbook.save(template_path)

            with patch.object(schedule, 'TEMPLATE_PATH', template_path):
                response = client.post('/export', json={
                    'people': [{'name': 'Alex', 'ranges': [{'start': '11:00', 'end': '11:30'}]}],
                    'room_data': {'11:00': {'off': 'Officer A', 'b': '2', 's': '1'}},
                    'is_pm': False,
                })

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response.mimetype or '',
        )
        self.assertIn('schedule-output.xlsx', response.headers['Content-Disposition'])

        exported = load_workbook(io.BytesIO(response.data))
        self.assertIn('Source Data', exported.sheetnames)
        self.assertEqual(exported['Source Data']['A1'].value, 'Alex, 11:00, 11:30')


if __name__ == '__main__':
    unittest.main()
