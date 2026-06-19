import os
import tempfile
import unittest
from typing import TypedDict
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import schedule
from schedule import build_xlsx


class TimeRange(TypedDict):
    start: str
    end: str


class Person(TypedDict):
    name: str
    ranges: list[TimeRange]


def make_person(name: str, ranges: list[TimeRange]) -> Person:
    return {'name': name, 'ranges': ranges}


class WorkbookTests(unittest.TestCase):
    def test_build_xlsx_writes_source_sheet_and_values(self):
        people = [
            make_person('Alex', [{'start': '11:00', 'end': '11:30'}]),
            make_person('Blake', [{'start': '14:00', 'end': '14:30'}]),
        ]
        room_data = {
            '11:00': {'off': 'Officer A', 'b': '3', 's': '3'},
            '14:00': {'off': 'Officer B', 'b': '2', 's': '1'},
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, 'template.xlsx')
            workbook = Workbook()
            active_sheet = workbook.active
            if active_sheet is not None:
                workbook.remove(active_sheet)
            workbook.create_sheet('AM')
            workbook.create_sheet('PM')
            workbook.save(template_path)

            with patch.object(schedule, 'TEMPLATE_PATH', template_path):
                buffer = build_xlsx(people, room_data)

        result = load_workbook(buffer)
        self.assertIn('Source Data', result.sheetnames)
        self.assertIn('AM', result.sheetnames)
        self.assertIn('PM', result.sheetnames)

        source = result['Source Data']
        self.assertEqual(source['A1'].value, 'Alex, 11:00, 11:30')
        self.assertEqual(source['A2'].value, 'Blake, 14:00, 14:30')
        self.assertEqual(source['A4'].value, '11:00, Officer A, 3, 3')

    def test_build_xlsx_escapes_formula_like_names_in_source_sheet(self):
        people = [
            {'name': '=cmd|\' /C calc\'!A0', 'ranges': [{'start': '11:00', 'end': '11:30'}]},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = os.path.join(tmpdir, 'template.xlsx')
            workbook = Workbook()
            active_sheet = workbook.active
            if active_sheet is not None:
                workbook.remove(active_sheet)
            workbook.create_sheet('AM')
            workbook.create_sheet('PM')
            workbook.save(template_path)

            with patch.object(schedule, 'TEMPLATE_PATH', template_path):
                buffer = build_xlsx(people, {})

        exported = load_workbook(buffer)
        self.assertEqual(exported['Source Data']['A1'].value, "'=cmd|' /C calc'!A0, 11:00, 11:30")

    def test_build_xlsx_raises_when_template_is_missing(self):
        with patch.object(schedule, 'TEMPLATE_PATH', '/definitely/missing/template.xlsx'):
            with self.assertRaises(FileNotFoundError):
                build_xlsx(
                    [make_person('Alex', [{'start': '11:00', 'end': '11:30'}])],
                    {},
                )


if __name__ == '__main__':
    unittest.main()
